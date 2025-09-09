import datetime
import logging
import os
from urllib.parse import quote
import mimetypes

from flask import request, make_response, send_file, session
import pandas as pd
from io import BytesIO

from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from jd import app, db
from jd.models.tg_account import TgAccount
from jd.models.tg_group import TgGroup
from jd.models.tg_group_chat_history import TgGroupChatHistory
from jd.models.tg_group_user_info import TgGroupUserInfo
from jd.models.tg_document_info import TgDocumentInfo
from jd.services.role_service.role import ROLE_SUPER_ADMIN, RoleService
from jd.views import get_or_exception
from jd.views.api import api
from flask import jsonify

logger = logging.getLogger(__name__)

def get_mime_type_from_path(file_path, file_ext=None):
    """根据文件路径或扩展名获取MIME类型"""
    # 首先尝试通过完整路径获取
    if file_path:
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type:
            return mime_type
    
    # 如果路径无法识别，尝试通过扩展名
    if file_ext:
        # 确保扩展名以点开头
        if not file_ext.startswith('.'):
            file_ext = '.' + file_ext
        mime_type, _ = mimetypes.guess_type('dummy' + file_ext)
        if mime_type:
            return mime_type
    
    # 返回默认值
    return 'application/octet-stream'

def is_image_mime_type(mime_type):
    """判断MIME类型是否为图片"""
    return mime_type and mime_type.startswith('image/')


def is_sticker_file(filename_origin, mime_type=None, filepath=None):
    """
    判断文件是否为贴图表情(sticker)
    优先使用原始文件名进行判断
    """
    # 优先使用原始文件名判断
    if filename_origin:
        filename_lower = filename_origin.lower()
        # 常见的sticker文件名特征
        sticker_patterns = [
            'sticker',
            'animatedsticker', 
            'videosticker',
            '.tgs',  # Telegram动画贴图格式
            '.webm', # 视频贴图
            '.webp', # 静态贴图
        ]
        
        for pattern in sticker_patterns:
            if pattern in filename_lower:
                return True
    
    # 如果没有原始文件名，则使用MIME类型和文件路径进行判断
    if mime_type:
        # 常见的sticker MIME类型
        sticker_mimes = [
            'application/x-tgsticker',  # Telegram sticker
            'image/webp',               # WebP sticker
            'video/webm'                # 视频 sticker
        ]
        if mime_type in sticker_mimes:
            return True
    
    # 最后使用文件路径判断(向后兼容)
    if filepath:
        filepath_lower = filepath.lower()
        if 'sticker' in filepath_lower or filepath_lower.endswith('.tgs') or filepath_lower.endswith('.webp'):
            return True
    
    return False


def get_sticker_display_info(filename_origin, mime_type=None, filepath=None):
    """
    获取贴图的显示信息
    返回包含显示图标和类型的信息
    """
    if not is_sticker_file(filename_origin, mime_type, filepath):
        return None
    
    # 根据文件类型返回不同的emoji图标
    if filename_origin:
        filename_lower = filename_origin.lower()
        if 'animated' in filename_lower or '.tgs' in filename_lower:
            return {
                'is_sticker': True,
                'sticker_type': 'animated',
                'display_icon': '🎭',  # 动画表情emoji
                'display_text': '【动画表情】'
            }
        elif '.webm' in filename_lower or 'video' in filename_lower:
            return {
                'is_sticker': True, 
                'sticker_type': 'video',
                'display_icon': '🎥',  # 视频表情emoji
                'display_text': '【动画表情】'
            }
    
    # 默认静态贴图
    return {
        'is_sticker': True,
        'sticker_type': 'static',
        'display_icon': '😀',  # 静态表情emoji 
        'display_text': '【动画表情】'
    }


def process_document_info(row_data):
    """
    处理文档信息的公共函数
    :param row_data: 包含聊天记录和文档信息的数据行
    :return: 处理后的文档列表
    """
    # 从连接结果中提取数据
    r = row_data[0]  # TgGroupChatHistory 对象
    doc_filename = row_data[1] if row_data[1] else None  # filename_origin
    doc_ext_name = row_data[2] if row_data[2] else None  # file_ext_name  
    doc_mime_type = row_data[3] if row_data[3] else None  # mime_type
    doc_filepath = row_data[4] if row_data[4] else None  # filepath
    doc_thumb_path = row_data[5] if row_data[5] else None  # video_thumb_path
    doc_hash = row_data[6] if row_data[6] else None  # file_hash
    doc_size = row_data[7] if row_data[7] else 0  # file_size
    doc_peer_id = row_data[8] if row_data[8] else None  # peer_id
    
    documents = []
    
    if doc_filename:  # 优先使用数据库中的详细文档信息（有filename就处理，filepath可能为空）
        # 检查是否为sticker并获取显示信息
        sticker_info = get_sticker_display_info(
            doc_filename.strip() if doc_filename else '',
            doc_mime_type.strip() if doc_mime_type else None,
            doc_filepath.strip() if doc_filepath else ''
        )
        
        doc_info = {
            'path': doc_filepath.strip() if doc_filepath else '',
            'filename_origin': doc_filename.strip() if doc_filename else '',
            'ext': doc_ext_name.strip() if doc_ext_name else '',
            'mime_type': doc_mime_type.strip() if doc_mime_type else '',
            'file_hash': doc_hash.strip() if doc_hash else '',
            'file_size': doc_size if doc_size else 0,
            'video_thumb_path': doc_thumb_path.strip() if doc_thumb_path else ''
        }
        
        # 如果是 sticker，添加 sticker 相关信息
        if sticker_info:
            doc_info.update(sticker_info)
        
        documents.append(doc_info)
    elif r.document_path:  # 如果没有详细信息，使用原有逻辑
        doc_ext = r.document_ext.strip() if r.document_ext else ''
        # 如果没有扩展名，尝试从路径中提取
        if not doc_ext and r.document_path:
            path_parts = r.document_path.split('.')
            if len(path_parts) > 1:
                doc_ext = path_parts[-1]
        
        # 获取MIME类型
        mime_type = get_mime_type_from_path(r.document_path, doc_ext)
        
        # 检查是否为sticker(使用路径判断)
        sticker_info = get_sticker_display_info(
            '',  # 没有原始文件名
            mime_type,
            r.document_path.strip()
        )
        
        doc_info = {
            'path': r.document_path.strip(),
            'ext': doc_ext,
            'mime_type': mime_type,
            'filename_origin': '',
            'file_hash': '',
            'file_size': 0,
            'video_thumb_path': ''
        }
        
        # 如果是 sticker，添加 sticker 相关信息
        if sticker_info:
            doc_info.update(sticker_info)
        
        documents.append(doc_info)
    
    return documents


@api.route('/tg/chat_room/history', methods=['GET'])
def tg_chat_room_history():
    try:
        args = request.args
        page = get_or_exception('page', args, 'int', 1)
        page_size = get_or_exception('page_size', args, 'int', 50)
        search_content = get_or_exception('search_content', args, 'str', '')
        start_date = get_or_exception('start_date', args, 'str', '')
        end_date = get_or_exception('end_date', args, 'str', '')
        message_id = get_or_exception('message_id', args, 'int', 0)
        reply_to_msg_id = get_or_exception('reply_to_msg_id', args, 'int', 0)
        search_chat_id_list = args.getlist('search_group_id')
        search_user_id_list = args.getlist('search_user_id')
        search_account_id_list = args.getlist('search_account_id')

        rows, total_records = fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list,
                                                          search_content, page, page_size, search_account_id_list,
                                                          message_id, reply_to_msg_id)
        
        total_pages = (total_records + page_size - 1) // page_size
        chat_room = TgGroup.query.filter_by(status=TgGroup.StatusType.JOIN_SUCCESS).all()
        group_list = [{'chat_id': c.chat_id, 'group_name': f'{c.name}-{c.title}'} for c in chat_room]
        chat_room = {r.chat_id: r.title for r in chat_room}
        data = []
        for row in rows:
            # 从连接结果中提取数据
            r = row[0]  # TgGroupChatHistory 对象
            doc_filename = row[1] if row[1] else None  # filename_origin
            doc_ext_name = row[2] if row[2] else None  # file_ext_name  
            doc_mime_type = row[3] if row[3] else None  # mime_type
            doc_filepath = row[4] if row[4] else None  # filepath
            doc_thumb_path = row[5] if row[5] else None  # video_thumb_path
            doc_hash = row[6] if row[6] else None  # file_hash
            doc_size = row[7] if row[7] else 0  # file_size
            doc_peer_id = row[8] if row[8] else None  # peer_id
            
            group_name = chat_room.get(r.chat_id, '')
            reply_to_msg_id = int(r.reply_to_msg_id) if r.reply_to_msg_id and r.reply_to_msg_id.isdigit() else 0
            print(f'chat_id:{r.chat_id}, path:{r.document_path}, enhanced_path:{doc_filepath}')
            
            # 处理document路径和类型信息，优先使用数据库中的详细信息
            documents = []
            if doc_filename:  # 优先使用数据库中的详细文档信息（有filename就处理，filepath可能为空）
                doc_info = {
                    'path': doc_filepath.strip() if doc_filepath else '',
                    'filename_origin': doc_filename.strip() if doc_filename else '',
                    'ext': doc_ext_name.strip() if doc_ext_name else '',
                    'mime_type': doc_mime_type.strip() if doc_mime_type else '',
                    'file_hash': doc_hash.strip() if doc_hash else '',
                    'file_size': doc_size if doc_size else 0,
                    'video_thumb_path': doc_thumb_path.strip() if doc_thumb_path else ''
                }
                documents.append(doc_info)
            elif r.document_path:  # 如果没有详细信息，使用原有逻辑
                document_paths = [r.document_path] if r.document_path else []
                document_exts = [r.document_ext] if r.document_ext else []
                
                # 组合document信息，包含路径和类型
                for i, path in enumerate(document_paths):
                    if path.strip():  # 确保路径不为空
                        doc_ext = document_exts[i].strip() if i < len(document_exts) and document_exts[i] else ''
                        # 如果没有扩展名，尝试从路径中提取
                        if not doc_ext and path:
                            path_parts = path.split('.')
                            if len(path_parts) > 1:
                                doc_ext = path_parts[-1]
                        
                        # 获取MIME类型
                        mime_type = get_mime_type_from_path(path, doc_ext)
                        
                        doc_info = {
                            'path': path.strip(),
                            'ext': doc_ext,
                            'mime_type': mime_type,
                            'filename_origin': '',
                            'file_hash': '',
                            'file_size': 0,
                            'video_thumb_path': ''
                        }
                        documents.append(doc_info)

            data.append({
                'id': r.id,
                'group_name': group_name,
                'message': r.message,
                'nickname': r.nickname,
                'postal_time': r.postal_time,
                'username': r.username,
                'user_id': r.user_id,
                'photo_paths': [r.photo_path] if r.photo_path else [],
                'document_paths': [r.document_path] if r.document_path else [],  # 保持向后兼容
                'documents': documents,  # 新增：包含类型信息的文档列表
                'reply_to_msg_id': reply_to_msg_id,
                'message_ids': r.message_id,
                'chat_id': r.chat_id
            })
        tg_group_user_info = TgGroupUserInfo.query.filter(TgGroupUserInfo.username != '').all()
        unique_users = {}
        for t in tg_group_user_info:
            unique_users[t.username] = {
                'user_id': t.user_id,
                'chat_id': t.chat_id,
                'nickname': t.nickname,
                'desc': t.desc,
                'photo': t.photo,
                'username': f'{t.username}-{t.nickname}'
            }

        # Convert the dictionary values back into a list
        group_user_list = list(unique_users.values())

        tg_accounts = TgAccount.query.filter(TgAccount.status == TgAccount.StatusType.JOIN_SUCCESS).all()
        tg_accounts_list = [{'account_id': t.id, 'username': t.username} for t in tg_accounts]

        # Redirect to JSON endpoint - the Vue frontend should use /history/json instead
        return jsonify({
            'err_code': 302,
            'err_msg': 'Redirect to Vue frontend',
            'payload': {'redirect': '/tg/chat-history'}
        }), 302
                               
    except Exception as e:
        logger.error(f'获取聊天记录页面失败: {e}', exc_info=True)
        return f"Internal Server Error: {str(e)}", 500


@api.route('/tg/chat_room/history/json', methods=['GET'])
def tg_chat_room_history_json():
    """返回JSON格式的聊天记录数据"""
    try:
        args = request.args
        page = get_or_exception('page', args, 'int', 1)
        page_size = get_or_exception('page_size', args, 'int', 20)
        search_content = get_or_exception('search_content', args, 'str', '')
        start_date = get_or_exception('start_date', args, 'str', '')
        end_date = get_or_exception('end_date', args, 'str', '')
        message_id = get_or_exception('message_id', args, 'int', 0)
        reply_to_msg_id = get_or_exception('reply_to_msg_id', args, 'int', 0)
        
        # 支持单个群组ID参数和多个群组ID参数
        group_id = get_or_exception('group_id', args, 'str', '')
        search_chat_id_list = args.getlist('search_group_id')
        if group_id:
            search_chat_id_list = [group_id]
        
        # 支持查看所有历史记录的选项
        show_all = get_or_exception('show_all', args, 'str', '')
        
        search_user_id_list = args.getlist('search_user_id')
        search_account_id_list = args.getlist('search_account_id')

        rows, total_records = fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list,
                                                          search_content, page, page_size, search_account_id_list,
                                                          message_id, reply_to_msg_id, show_all)
        total_pages = (total_records + page_size - 1) // page_size
        
        # 获取群组信息
        chat_room = TgGroup.query.filter_by(status=TgGroup.StatusType.JOIN_SUCCESS).all()
        group_list = [{'chat_id': c.chat_id, 'group_name': f'{c.name}-{c.title}'} for c in chat_room]
        chat_room_dict = {r.chat_id: r.title for r in chat_room}
        
        # 获取所有用户头像信息和关注状态
        user_avatars = {}
        user_focus_status = {}
        if rows:
            chat_ids = list(set([r[0].chat_id for r in rows]))
            user_ids = list(set([r[0].user_id for r in rows]))
            user_infos = TgGroupUserInfo.query.filter(
                TgGroupUserInfo.chat_id.in_(chat_ids),
                TgGroupUserInfo.user_id.in_(user_ids)
            ).all()
            
            for user_info in user_infos:
                key = f"{user_info.chat_id}_{user_info.user_id}"
                user_avatars[key] = user_info.avatar_path or user_info.photo
                user_focus_status[key] = bool(user_info.is_key_focus) if hasattr(user_info, 'is_key_focus') else False
        
        # 处理聊天记录数据
        data = []
        for row in rows:
            # 从连接结果中提取数据
            r = row[0]  # TgGroupChatHistory 对象
            doc_filename = row[1] if row[1] else None  # filename_origin
            doc_ext_name = row[2] if row[2] else None  # file_ext_name  
            doc_mime_type = row[3] if row[3] else None  # mime_type
            doc_filepath = row[4] if row[4] else None  # filepath
            doc_thumb_path = row[5] if row[5] else None  # video_thumb_path
            doc_hash = row[6] if row[6] else None  # file_hash
            doc_size = row[7] if row[7] else 0  # file_size
            doc_peer_id = row[8] if row[8] else None  # peer_id
            
            group_name = chat_room_dict.get(r.chat_id, '')
            avatar_key = f"{r.chat_id}_{r.user_id}"
            user_avatar = user_avatars.get(avatar_key, '')
            is_key_focus = user_focus_status.get(avatar_key, False)
            
            # 使用公共函数处理文档信息
            documents = process_document_info(row)

            data.append({
                'id': r.id,
                'group_name': group_name,
                'message': r.message,
                'nickname': r.nickname,
                'postal_time': r.postal_time.strftime('%Y-%m-%d %H:%M:%S') if r.postal_time else '',
                'username': r.username,
                'user_id': r.user_id,
                'user_avatar': user_avatar,
                'is_key_focus': is_key_focus,
                'photo_paths': [r.photo_path] if r.photo_path else [],
                'document_paths': [r.document_path] if r.document_path else [],  # 保持向后兼容
                'documents': documents,  # 新增：包含类型信息的文档列表
                'reply_to_msg_id': r.reply_to_msg_id or 0,
                'message_ids': r.message_id,
                'chat_id': r.chat_id
            })
        
        # 获取用户信息
        tg_group_user_info = TgGroupUserInfo.query.filter(TgGroupUserInfo.username != '').all()
        unique_users = {}
        for t in tg_group_user_info:
            unique_users[t.username] = {
                'user_id': t.user_id,
                'chat_id': t.chat_id,
                'nickname': t.nickname,
                'desc': t.desc,
                'photo': t.photo,
                'username': f'{t.username}-{t.nickname}'
            }
        group_user_list = list(unique_users.values())
        
        # 获取TG账户信息
        tg_accounts = TgAccount.query.filter(TgAccount.status == TgAccount.StatusType.JOIN_SUCCESS).all()
        tg_accounts_list = [{'account_id': t.id, 'username': t.username} for t in tg_accounts]

        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'data': data,
                'total_pages': total_pages,
                'current_page': page,
                'total_records': total_records,
                'group_list': group_list,
                'group_user_list': group_user_list,
                'tg_accounts': tg_accounts_list
            }
        })
        
    except Exception as e:
        logger.error(f'获取聊天记录失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/chat_room/history/by_group/<group_id>', methods=['GET'])
def tg_chat_room_history_by_group(group_id):
    """根据群组ID返回该群组的聊天记录"""
    try:
        args = request.args
        page = get_or_exception('page', args, 'int', 1)
        page_size = get_or_exception('page_size', args, 'int', 20)
        search_content = get_or_exception('search_content', args, 'str', '')
        start_date = get_or_exception('start_date', args, 'str', '')
        end_date = get_or_exception('end_date', args, 'str', '')
        message_id = get_or_exception('message_id', args, 'int', 0)
        reply_to_msg_id = get_or_exception('reply_to_msg_id', args, 'int', 0)
        
        # 直接使用传入的群组ID
        search_chat_id_list = [group_id]
        search_user_id_list = args.getlist('search_user_id')
        search_account_id_list = args.getlist('search_account_id')
        
        # 支持查看所有历史记录的选项
        show_all = get_or_exception('show_all', args, 'str', '')

        rows, total_records = fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list,
                                                          search_content, page, page_size, search_account_id_list,
                                                          message_id, reply_to_msg_id, show_all)
        total_pages = (total_records + page_size - 1) // page_size
        
        # 获取群组信息
        chat_room = TgGroup.query.filter_by(chat_id=group_id, status=TgGroup.StatusType.JOIN_SUCCESS).first()
        if not chat_room:
            return jsonify({
                'err_code': 1,
                'err_msg': '群组不存在或未成功加入',
                'payload': {}
            }), 404
        
        group_info = {'chat_id': chat_room.chat_id, 'group_name': f'{chat_room.name}-{chat_room.title}'}
        
        # 获取所有用户头像信息和关注状态
        user_avatars = {}
        user_focus_status = {}
        if rows:
            user_infos = TgGroupUserInfo.query.filter(
                TgGroupUserInfo.chat_id == group_id,
                TgGroupUserInfo.user_id.in_([r[0].user_id for r in rows])
            ).all()
            
            for user_info in user_infos:
                key = f"{user_info.chat_id}_{user_info.user_id}"
                user_avatars[key] = user_info.avatar_path or user_info.photo
                user_focus_status[key] = bool(user_info.is_key_focus) if hasattr(user_info, 'is_key_focus') else False
        
        # 处理聊天记录数据
        data = []
        for row in rows:
            # 从连接结果中提取数据
            r = row[0]  # TgGroupChatHistory 对象
            doc_filename = row[1] if row[1] else None  # filename_origin
            doc_ext_name = row[2] if row[2] else None  # file_ext_name  
            doc_mime_type = row[3] if row[3] else None  # mime_type
            doc_filepath = row[4] if row[4] else None  # filepath
            doc_thumb_path = row[5] if row[5] else None  # video_thumb_path
            doc_hash = row[6] if row[6] else None  # file_hash
            doc_size = row[7] if row[7] else 0  # file_size
            doc_peer_id = row[8] if row[8] else None  # peer_id
            
            avatar_key = f"{r.chat_id}_{r.user_id}"
            user_avatar = user_avatars.get(avatar_key, '')
            is_key_focus = user_focus_status.get(avatar_key, False)
            
            # 处理document路径和类型信息，优先使用数据库中的详细信息
            documents = []
            if doc_filename:  # 优先使用数据库中的详细文档信息（有filename就处理，filepath可能为空）
                # 检查是否为sticker并获取显示信息
                sticker_info = get_sticker_display_info(
                    doc_filename.strip() if doc_filename else '',
                    doc_mime_type.strip() if doc_mime_type else None,
                    doc_filepath.strip() if doc_filepath else ''
                )
                
                doc_info = {
                    'path': doc_filepath.strip() if doc_filepath else '',
                    'filename_origin': doc_filename.strip() if doc_filename else '',
                    'ext': doc_ext_name.strip() if doc_ext_name else '',
                    'mime_type': doc_mime_type.strip() if doc_mime_type else '',
                    'file_hash': doc_hash.strip() if doc_hash else '',
                    'file_size': doc_size if doc_size else 0,
                    'video_thumb_path': doc_thumb_path.strip() if doc_thumb_path else ''
                }
                
                # 如果是 sticker，添加 sticker 相关信息
                if sticker_info:
                    doc_info.update(sticker_info)
                
                documents.append(doc_info)
            elif r.document_path:  # 如果没有详细信息，使用原有逻辑
                doc_ext = r.document_ext.strip() if r.document_ext else ''
                # 如果没有扩展名，尝试从路径中提取
                if not doc_ext and r.document_path:
                    path_parts = r.document_path.split('.')
                    if len(path_parts) > 1:
                        doc_ext = path_parts[-1]
                
                # 获取MIME类型
                mime_type = get_mime_type_from_path(r.document_path, doc_ext)
                
                doc_info = {
                    'path': r.document_path.strip(),
                    'ext': doc_ext,
                    'mime_type': mime_type,
                    'filename_origin': '',
                    'file_hash': '',
                    'file_size': 0,
                    'video_thumb_path': ''
                }
                documents.append(doc_info)

            data.append({
                'id': r.id,
                'message_id': r.message_id,
                'group_name': chat_room.title,
                'message': r.message,
                'nickname': r.nickname,
                'postal_time': r.postal_time.strftime('%Y-%m-%d %H:%M:%S') if r.postal_time else '',
                'username': r.username,
                'user_id': r.user_id,
                'user_avatar': user_avatar,
                'is_key_focus': is_key_focus,
                'photo_paths': [r.photo_path] if r.photo_path else [],
                'document_paths': [r.document_path] if r.document_path else [],  # 保持向后兼容
                'documents': documents,  # 新增：包含类型信息的文档列表
                'reply_to_msg_id': r.reply_to_msg_id or 0,
                'message_ids': r.message_id,
                'chat_id': r.chat_id
            })

        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'data': data,
                'total_pages': total_pages,
                'current_page': page,
                'total_records': total_records,
                'group_info': group_info
            }
        })
        
    except Exception as e:
        logger.error(f'根据群组ID获取聊天记录失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/chat_room/history/by_private_chat/<chat_id>', methods=['GET'])
def tg_chat_room_history_by_private_chat(chat_id):
    """根据私人聊天ID返回该私聊的聊天记录"""
    try:
        args = request.args
        page = get_or_exception('page', args, 'int', 1)
        page_size = get_or_exception('page_size', args, 'int', 20)
        search_content = get_or_exception('search_content', args, 'str', '')
        start_date = get_or_exception('start_date', args, 'str', '')
        end_date = get_or_exception('end_date', args, 'str', '')
        message_id = get_or_exception('message_id', args, 'int', 0)
        reply_to_msg_id = get_or_exception('reply_to_msg_id', args, 'int', 0)
        
        # 私人聊天使用chat_id进行查询
        search_chat_id_list = [chat_id]
        search_user_id_list = []
        search_account_id_list = args.getlist('search_account_id')
        
        # 支持查看所有历史记录的选项
        show_all = get_or_exception('show_all', args, 'str', '')
        rows, total_records = fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list,
                                                         search_content, page, page_size, search_account_id_list,
                                                         message_id, reply_to_msg_id, show_all)

        # 验证是否有数据
        if not rows:
            return jsonify({
                'err_code': 1,
                'err_msg': '该私人聊天暂无记录或不存在',
                'payload': {}
            }), 404

        # 获取私人聊天的基本信息
        first_row = rows[0] if rows else None
        chat_info = {
            'chat_id': chat_id, 
            'group_name': f"与 {first_row.nickname or first_row.username or '未知用户'} 的私聊" if first_row else f"私人聊天 {chat_id}"
        }
        
        # 获取用户头像信息
        user_avatars = {}
        if rows:
            user_infos = TgGroupUserInfo.query.filter(
                TgGroupUserInfo.chat_id == chat_id,
                TgGroupUserInfo.user_id.in_([r[0].user_id for r in rows])
            ).all()
            
            for user_info in user_infos:
                key = f"{user_info.chat_id}_{user_info.user_id}"
                user_avatars[key] = user_info.avatar_path or user_info.photo
        
        # 处理聊天记录数据
        data = []
        for row in rows:
            # 从连接结果中提取数据
            r = row[0]  # TgGroupChatHistory 对象
            doc_filename = row[1] if row[1] else None  # filename_origin
            doc_ext_name = row[2] if row[2] else None  # file_ext_name  
            doc_mime_type = row[3] if row[3] else None  # mime_type
            doc_filepath = row[4] if row[4] else None  # filepath
            doc_thumb_path = row[5] if row[5] else None  # video_thumb_path
            doc_hash = row[6] if row[6] else None  # file_hash
            doc_size = row[7] if row[7] else 0  # file_size
            doc_peer_id = row[8] if row[8] else None  # peer_id
            
            avatar_key = f"{r.chat_id}_{r.user_id}"
            user_avatar = user_avatars.get(avatar_key, '')
            
            # 使用公共函数处理文档信息
            documents = process_document_info(row)

            data.append({
                'id': r.id,
                'group_name': chat_info['group_name'],
                'chat_id': r.chat_id,
                'user_id': r.user_id,
                'message': r.message,
                'nickname': r.nickname,
                'username': r.username,
                'postal_time': r.postal_time.strftime('%Y-%m-%d %H:%M:%S') if r.postal_time else '',
                'photo_paths': r.photo_path.split(',') if r.photo_path else [],
                'document_paths': [r.document_path] if r.document_path else [],  # 保持向后兼容
                'documents': documents,  # 新增：包含类型信息的文档列表
                'reply_to_msg_id': r.reply_to_msg_id or 0,
                'message_ids': r.message_id or '',
                'user_avatar': user_avatar,
                'is_key_focus': False  # 私人聊天默认不设置重点关注
            })

        return jsonify({
            'err_code': 0,
            'err_msg': 'success',
            'payload': {
                'data': data,
                'total_pages': (total_records + page_size - 1) // page_size,
                'current_page': page,
                'total_records': total_records,
                'group_list': [chat_info]
            }
        })
        
    except Exception as e:
        logger.error(f'根据私人聊天ID获取聊天记录失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


def fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list, search_content,
                                page=None,
                                page_size=None, search_account_id_list=None, message_id=0, reply_to_msg_id=0, show_all=''):
    # 直接查询所有消息，并左连接文档信息表
    query = db.session.query(
        TgGroupChatHistory,
        TgDocumentInfo.filename_origin,
        TgDocumentInfo.file_ext_name,
        TgDocumentInfo.mime_type,
        TgDocumentInfo.filepath,
        TgDocumentInfo.video_thumb_path,
        TgDocumentInfo.file_hash,
        TgDocumentInfo.file_size,
        TgDocumentInfo.peer_id
    ).outerjoin(
        TgDocumentInfo,
        (TgGroupChatHistory.chat_id == TgDocumentInfo.chat_id) &
        (TgGroupChatHistory.message_id == TgDocumentInfo.message_id)
    )

    if start_date and end_date:
        f_start_date = start_date + ' 00:00:00'
        f_end_date = end_date + ' 23:59:59'
        query = query.filter(TgGroupChatHistory.postal_time.between(f_start_date, f_end_date))
    elif show_all != 'true':
        # 只有在不是查看全部记录时才应用默认时间限制
        now_time = datetime.datetime.now()
        start_time = (now_time - datetime.timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')
        end_time = now_time.strftime('%Y-%m-%d 23:59:59')
        query = query.filter(TgGroupChatHistory.postal_time.between(start_time, end_time))
    search_chat_id_list = [r for r in search_chat_id_list if r]
    if search_chat_id_list:
        query = query.filter(TgGroupChatHistory.chat_id.in_(search_chat_id_list))
    search_user_id_list = [r for r in search_user_id_list if r]
    if search_user_id_list:
        query = query.filter(TgGroupChatHistory.user_id.in_(search_user_id_list))
    if search_content:
        query = query.filter(TgGroupChatHistory.message.like(f'%{search_content}%'))
    search_account_id_list = [r for r in search_account_id_list if r]
    if search_account_id_list:
        tg_accounts = TgAccount.query.filter(TgAccount.id.in_(search_account_id_list)).all()
        user_id_list = [t.user_id for t in tg_accounts]
        his = TgGroupChatHistory.query.filter(TgGroupChatHistory.user_id.in_(user_id_list)).all()
        chat_id_list = [t.chat_id for t in his]
        query = query.filter(TgGroupChatHistory.chat_id.in_(chat_id_list))
    if reply_to_msg_id:
        query = query.filter(TgGroupChatHistory.message_id == reply_to_msg_id)
    if message_id:
        message = query.filter(TgGroupChatHistory.id == message_id).first()
        if message:
            start_time = message.postal_time - datetime.timedelta(hours=1)
            end_time = message.postal_time + datetime.timedelta(minutes=5)
            query = query.filter(TgGroupChatHistory.chat_id == message.chat_id,
                                 TgGroupChatHistory.postal_time >= start_time,
                                 TgGroupChatHistory.postal_time <= end_time)

    # 获取总数
    total_records = query.count()
    
    # 使用数据库层面分页，并按时间排序（从旧到新）
    query = query.order_by(TgGroupChatHistory.postal_time.asc())
    
    # 应用分页
    if page and page_size:
        query = query.offset((page - 1) * page_size).limit(page_size)
    
    rows = query.all()
        
    return rows, total_records


@api.route('/tg/chat_room/history/by_user_in_group/<group_id>/<user_id>', methods=['GET'])
def tg_chat_room_history_by_user_in_group(group_id, user_id):
    """根据用户数字ID和群组ID返回该用户在该群组的全部聊天记录"""
    try:
        args = request.args
        page = get_or_exception('page', args, 'int', 1)
        page_size = get_or_exception('page_size', args, 'int', 20)
        search_content = get_or_exception('search_content', args, 'str', '')
        start_date = get_or_exception('start_date', args, 'str', '')
        end_date = get_or_exception('end_date', args, 'str', '')
        
        # 验证参数
        if not group_id or not user_id:
            return jsonify({
                'err_code': 1,
                'err_msg': '群组ID和用户ID不能为空',
                'payload': {}
            }), 400
        
        # 固定搜索条件：特定群组和特定用户
        search_chat_id_list = [group_id]
        search_user_id_list = [user_id]
        search_account_id_list = []
        
        # 默认查看所有历史记录（不限制时间范围）
        show_all = 'true'

        rows, total_records = fetch_tg_group_chat_history(start_date, end_date, search_chat_id_list, search_user_id_list,
                                                          search_content, page, page_size, search_account_id_list,
                                                          0, 0, show_all)
        total_pages = (total_records + page_size - 1) // page_size
        
        # 获取群组信息
        chat_room = TgGroup.query.filter_by(chat_id=group_id, status=TgGroup.StatusType.JOIN_SUCCESS).first()
        if not chat_room:
            return jsonify({
                'err_code': 1,
                'err_msg': '群组不存在或未成功加入',
                'payload': {}
            }), 404
        
        # 获取用户信息
        user_info = TgGroupUserInfo.query.filter_by(chat_id=group_id, user_id=user_id).first()
        user_details = {
            'user_id': user_id,
            'username': user_info.username if user_info else '',
            'nickname': user_info.nickname if user_info else '',
            'avatar_path': user_info.avatar_path or user_info.photo if user_info else '',
            'is_key_focus': bool(user_info.is_key_focus) if user_info and hasattr(user_info, 'is_key_focus') else False
        }
        
        group_info = {
            'chat_id': chat_room.chat_id, 
            'group_name': f'{chat_room.name}-{chat_room.title}',
            'title': chat_room.title
        }
        
        # 处理聊天记录数据
        data = []
        for row in rows:
            # 从连接结果中提取数据
            r = row[0]  # TgGroupChatHistory 对象
            
            # 使用公共函数处理文档信息
            documents = process_document_info(row)

            data.append({
                'id': r.id,
                'message_id': r.message_id,
                'group_name': chat_room.title,
                'message': r.message,
                'nickname': r.nickname,
                'postal_time': r.postal_time.strftime('%Y-%m-%d %H:%M:%S') if r.postal_time else '',
                'username': r.username,
                'user_id': r.user_id,
                'user_avatar': user_details['avatar_path'],
                'is_key_focus': user_details['is_key_focus'],
                'photo_paths': [r.photo_path] if r.photo_path else [],
                'document_paths': [r.document_path] if r.document_path else [],  # 保持向后兼容
                'documents': documents,  # 新增：包含类型信息的文档列表
                'reply_to_msg_id': r.reply_to_msg_id or 0,
                'message_ids': r.message_id,
                'chat_id': r.chat_id
            })

        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'data': data,
                'total_pages': total_pages,
                'current_page': page,
                'total_records': total_records,
                'group_info': group_info,
                'user_info': user_details
            }
        })
        
    except Exception as e:
        logger.error(f'根据用户ID和群组ID获取聊天记录失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/chat_room/history/find_page/<group_id>/<int:message_id>', methods=['GET'])
def find_message_page(group_id, message_id):
    """根据消息ID查找消息所在的页数"""
    try:
        page_size = get_or_exception('page_size', request.args, 'int', 20)
        
        # 查找目标消息
        target_message = TgGroupChatHistory.query.filter_by(
            chat_id=group_id, 
            id=message_id
        ).first()
        
        if not target_message:
            return jsonify({
                'err_code': 1,
                'err_msg': '消息不存在',
                'payload': {}
            }), 404
        
        # 统计目标消息之前的消息数量（因为消息按时间正序排列，从旧到新）
        messages_before = TgGroupChatHistory.query.filter(
            TgGroupChatHistory.chat_id == group_id,
            TgGroupChatHistory.postal_time < target_message.postal_time
        ).count()
        
        # 计算页数（从1开始）
        page_number = (messages_before // page_size) + 1
        
        # 获取该页面的消息范围信息（保持与主API一致的排序）
        messages_in_page = TgGroupChatHistory.query.filter_by(chat_id=group_id)\
            .order_by(TgGroupChatHistory.postal_time.asc())\
            .offset((page_number - 1) * page_size)\
            .limit(page_size)\
            .all()
        
        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'page_number': page_number,
                'message_id': message_id,
                'message_time': target_message.postal_time.isoformat() if target_message.postal_time else None,
                'total_messages_before': messages_before,
                'page_size': page_size,
                'messages_in_page': len(messages_in_page)
            }
        })
        
    except Exception as e:
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500

@api.route('/tg/chat_room/history/download', methods=['GET'])
def tg_chat_room_history_download():
    args = request.args
    chat_id = get_or_exception('chat_id', args, 'str')
    start_time = get_or_exception('start_time', args, 'str', '')
    end_time = get_or_exception('end_time', args, 'str', '')
    get_photo = get_or_exception('get_photo', args, 'bool', False)
    
    # 验证必需的参数
    if not chat_id:
        return jsonify({
            'err_code': 1,
            'err_msg': '请选择要导出的群组',
            'payload': {}
        }), 400
    
    # 只导出指定群组的数据
    search_chat_id_list = [chat_id]
    search_user_id_list = []
    search_account_id_list = []
    search_content = ''

    logger.debug(f"导出群组聊天记录: chat_id={chat_id}, start_time={start_time}, end_time={end_time}, get_photo={get_photo}")
    
    rows, _ = fetch_tg_group_chat_history(start_time, end_time, search_chat_id_list, search_user_id_list,
                                          search_content, None, None, search_account_id_list)
    chat_room = TgGroup.query.filter_by(status=TgGroup.StatusType.JOIN_SUCCESS).all()
    chat_room = {r.chat_id: r.title for r in chat_room}
    data = []
    for row in rows:
        # 从连接结果中提取数据
        r = row[0]  # TgGroupChatHistory 对象
        # 文档信息在这里不需要处理，因为这个函数只是导出Excel
        
        group_name = chat_room.get(r.chat_id, '')
        data.append({
            '群组名称': group_name,
            # '内容': r.message,
            '昵称': r.nickname,
            '发布时间': r.postal_time.strftime('%Y-%m-%d %H:%M:%S'),
            '用户名': r.username,
            '用户ID': r.user_id,
            '图片': r.photo_path,  # 修正为photo_paths为photo_path
            '内容': r.message      # 修正为messages为message
        })

    # 创建DataFrame
    columns = ['群组名称', '昵称', '发布时间', '用户名', '用户ID', '内容', '图片']
    df = pd.DataFrame(data, columns=columns)

    # 将DataFrame保存到Excel文件
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # 获取Excel工作簿
    workbook = writer.book

    # 根据get_photo参数决定是否加载图片到Excel
    if get_photo:
        ws = workbook['Sheet1']
        for idx, row in df.iterrows():
            logger.debug(f"title:{row['群组名称']}, photo count:{len(row['图片'].split(','))}, paths:{row['图片']}, idx:{idx}")
            if pd.isna(row['图片']):
                continue
            if not row['图片']:
                continue

            i = 0
            for i, img_path in enumerate(row['图片'].split(',')):
                if not img_path:
                    continue
                img_path = os.path.join(app.static_folder, img_path)
                if not os.path.exists(img_path):
                    continue
                try:
                    img = Image(img_path)
                except Exception as e:
                    logger.error(f'图片加载错误：{img_path}, error:{e}')
                    continue
                # 调整图片大小
                img.width = 65
                img.height = 100
                # 将图片插入到对应的行
                cell = ws.cell(row=idx + 2, column=len(columns) + i)  # 假设图片放在最后一列
                cell.value = ''
                ws.add_image(img, cell.coordinate)
                column_letter = get_column_letter(len(columns) + i + 1)  # 获取列的字母标
                ws.column_dimensions[column_letter].width = 65 / 6  # 适当调整比例
            ws.row_dimensions[idx + 2].height = 100  # 适当调整比例
    writer.close()

    # 设置响应头
    output.seek(0)
    response = make_response(output.getvalue())
    file_name = 'chat_history.xlsx'
    response.headers["Content-Disposition"] = f"attachment; filename={quote(file_name)}; filename*=utf-8''{quote(file_name)}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return response


@api.route('/tg/chat_room/history/download_file', methods=['POST'])
def download_file_from_message():
    """根据chat_id和message_id重新下载聊天记录中的文件"""
    try:
        # 获取请求参数
        data = request.get_json()
        if not data:
            logger.warning("文件下载触发失败: 请求参数为空")
            return jsonify({
                'err_code': 1,
                'err_msg': '请求参数不能为空',
                'payload': {}
            }), 400
        
        chat_id = data.get('chat_id')
        message_id = data.get('message_id')
        
        if not chat_id or not message_id:
            logger.warning(f"文件下载触发失败: 参数不完整 - chat_id={chat_id}, message_id={message_id}")
            return jsonify({
                'err_code': 1,
                'err_msg': 'chat_id和message_id不能为空',
                'payload': {}
            }), 400
        
        # 记录下载触发
        logger.info(f"[文件下载触发] 异步模式 - chat_id={chat_id}, message_id={message_id}, 来源IP={request.remote_addr}")
        
        # 验证聊天记录是否存在
        chat_history = TgGroupChatHistory.query.filter_by(
            chat_id=str(chat_id),
            message_id=str(message_id)
        ).first()
        
        if not chat_history:
            logger.warning(f"[文件下载触发失败] 聊天记录不存在 - chat_id={chat_id}, message_id={message_id}")
            return jsonify({
                'err_code': 1,
                'err_msg': '聊天记录不存在',
                'payload': {}
            }), 404
        
        # 异步调用下载任务
        from jd.tasks.telegram.tg_download_file import download_file_by_message_task
        
        # 启动异步下载任务
        task_result = download_file_by_message_task.delay(chat_id, message_id)
        
        logger.info(f"[文件下载任务启动] 异步任务ID={task_result.id}, chat_id={chat_id}, message_id={message_id}")
        
        return jsonify({
            'err_code': 0,
            'err_msg': '文件下载任务已启动',
            'payload': {
                'task_id': task_result.id,
                'chat_id': chat_id,
                'message_id': message_id,
                'status': '下载中'
            }
        })
        
    except Exception as e:
        logger.error(f'启动文件下载任务失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': f'启动文件下载任务失败: {str(e)}',
            'payload': {}
        }), 500


@api.route('/tg/chat_room/history/download_file_sync', methods=['POST'])
async def download_file_from_message_sync():
    """同步版本：根据chat_id和message_id重新下载聊天记录中的文件"""
    try:
        # 获取请求参数
        data = request.get_json()
        if not data:
            logger.warning("文件下载触发失败: 请求参数为空 (同步模式)")
            return jsonify({
                'err_code': 1,
                'err_msg': '请求参数不能为空',
                'payload': {}
            }), 400
        
        chat_id = data.get('chat_id')
        message_id = data.get('message_id')
        
        if not chat_id or not message_id:
            logger.warning(f"文件下载触发失败: 参数不完整 (同步模式) - chat_id={chat_id}, message_id={message_id}")
            return jsonify({
                'err_code': 1,
                'err_msg': 'chat_id和message_id不能为空',
                'payload': {}
            }), 400
        
        # 记录下载触发
        logger.info(f"[文件下载触发] 同步模式 - chat_id={chat_id}, message_id={message_id}, 来源IP={request.remote_addr}")
        
        # 验证聊天记录是否存在
        chat_history = TgGroupChatHistory.query.filter_by(
            chat_id=str(chat_id),
            message_id=str(message_id)
        ).first()
        
        if not chat_history:
            logger.warning(f"[文件下载触发失败] 聊天记录不存在 (同步模式) - chat_id={chat_id}, message_id={message_id}")
            return jsonify({
                'err_code': 1,
                'err_msg': '聊天记录不存在',
                'payload': {}
            }), 404
        
        # 直接调用下载方法（同步版本，适用于测试）
        from jd.jobs.tg_file_info import TgFileInfoManager
        from jd.services.spider.tg import TgService
        
        logger.info(f"[文件下载开始] 同步模式 - 初始化TG服务, chat_id={chat_id}, message_id={message_id}")
        
        # 初始化TG服务
        tg_service = await TgService.init_tg('web')
        if not tg_service:
            logger.error(f"[文件下载失败] TG服务初始化失败 - chat_id={chat_id}, message_id={message_id}")
            return jsonify({
                'err_code': 1,
                'err_msg': 'Telegram服务初始化失败',
                'payload': {}
            }), 500
        
        try:
            # 设置文件保存路径
            document_path = os.path.join(app.static_folder, 'document')
            image_path = os.path.join(app.static_folder, 'images')
            
            # 确保目录存在
            os.makedirs(document_path, exist_ok=True)
            os.makedirs(image_path, exist_ok=True)
            
            # 调用下载方法
            import asyncio
            async def download_file():
                return await TgFileInfoManager.download_file_by_message(
                    client=tg_service.client,
                    chat_id=chat_id,
                    message_id=message_id,
                    document_path=document_path,
                    image_path=image_path
                )
            
            # 在新的事件循环中运行异步方法
            logger.info(f"[文件下载执行] 同步模式 - 开始执行下载, chat_id={chat_id}, message_id={message_id}")
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(download_file())
                logger.info(f"[文件下载完成] 同步模式 - success={result['success']}, message={result['message']}, chat_id={chat_id}, message_id={message_id}")
            finally:
                loop.close()
            
            return jsonify({
                'err_code': 0 if result['success'] else 1,
                'err_msg': result['message'],
                'payload': {
                    'chat_id': chat_id,
                    'message_id': message_id,
                    'file_type': result['file_type'],
                    'file_info': result['file_info']
                }
            })
            
        finally:
            # 清理TG服务
            if tg_service:
                tg_service.close_client()
        
    except Exception as e:
        logger.error(f'文件下载失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': f'文件下载失败: {str(e)}',
            'payload': {}
        }), 500


@api.route('/tg/chat_room/history/fetch_new', methods=['POST'])
def fetch_new_chat_history():
    """触发获取新的Telegram聊天历史记录"""
    try:
        from jd.tasks.first.tg_history_job import fetch_tg_history_job
        
        # 使用Celery异步执行任务
        task = fetch_tg_history_job.delay()
        
        return jsonify({
            'err_code': 0,
            'err_msg': '聊天历史获取任务已启动',
            'payload': {
                'task_id': task.id
            }
        })
        
    except Exception as e:
        logger.error(f'启动聊天历史获取任务失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': f'启动任务失败: {str(e)}',
            'payload': {}
        }), 500


@api.route('/tg/document/filename/<chat_id>/<message_id>', methods=['GET'])
def get_document_filename(chat_id, message_id):
    """获取文档的原始文件名"""
    try:
        # 验证参数
        if not chat_id or not message_id:
            return jsonify({
                'err_code': 1,
                'err_msg': 'chat_id和message_id不能为空',
                'payload': {}
            }), 400
        
        # 查询文档信息
        document_info = TgDocumentInfo.query.filter_by(
            chat_id=str(chat_id),
            message_id=str(message_id)
        ).first()
        
        if not document_info:
            return jsonify({
                'err_code': 1,
                'err_msg': '文档信息不存在',
                'payload': {}
            }), 404
        
        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'chat_id': chat_id,
                'message_id': message_id,
                'filename_origin': document_info.filename_origin,
                'file_ext_name': document_info.file_ext_name,
                'mime_type': document_info.mime_type,
                'filepath': document_info.filepath,
                'file_size': document_info.file_size,
                'file_hash': document_info.file_hash,
                'video_thumb_path': document_info.video_thumb_path,
                'created_at': document_info.created_at.strftime('%Y-%m-%d %H:%M:%S') if document_info.created_at else '',
                'updated_at': document_info.updated_at.strftime('%Y-%m-%d %H:%M:%S') if document_info.updated_at else ''
            }
        })
        
    except Exception as e:
        logger.error(f'获取文档信息失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/document/info/batch', methods=['POST'])
def get_documents_info_batch():
    """批量获取文档信息"""
    try:
        data = request.get_json()
        if not data or 'documents' not in data:
            return jsonify({
                'err_code': 1,
                'err_msg': '请求参数不能为空',
                'payload': {}
            }), 400
        
        documents_list = data['documents']
        if not isinstance(documents_list, list):
            return jsonify({
                'err_code': 1,
                'err_msg': 'documents参数必须是数组',
                'payload': {}
            }), 400
        
        # 构建查询条件
        conditions = []
        for doc in documents_list:
            if 'chat_id' in doc and 'message_id' in doc:
                conditions.append(
                    (TgDocumentInfo.chat_id == str(doc['chat_id'])) &
                    (TgDocumentInfo.message_id == str(doc['message_id']))
                )
        
        if not conditions:
            return jsonify({
                'err_code': 1,
                'err_msg': '没有有效的查询条件',
                'payload': {'documents': []}
            }), 400
        
        # 执行查询
        from sqlalchemy import or_
        document_infos = TgDocumentInfo.query.filter(or_(*conditions)).all()
        
        # 构建返回结果
        result_documents = []
        for doc_info in document_infos:
            result_documents.append({
                'chat_id': doc_info.chat_id,
                'message_id': doc_info.message_id,
                'filename_origin': doc_info.filename_origin,
                'file_ext_name': doc_info.file_ext_name,
                'mime_type': doc_info.mime_type,
                'filepath': doc_info.filepath,
                'file_size': doc_info.file_size,
                'file_hash': doc_info.file_hash,
                'video_thumb_path': doc_info.video_thumb_path,
                'created_at': doc_info.created_at.strftime('%Y-%m-%d %H:%M:%S') if doc_info.created_at else '',
                'updated_at': doc_info.updated_at.strftime('%Y-%m-%d %H:%M:%S') if doc_info.updated_at else ''
            })
        
        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'documents': result_documents,
                'total': len(result_documents)
            }
        })
        
    except Exception as e:
        logger.error(f'批量获取文档信息失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/document/filename_by_path', methods=['POST'])
def get_document_filename_by_path():
    """根据文件路径获取原始文件名"""
    try:
        data = request.get_json()
        if not data or 'filepath' not in data:
            return jsonify({
                'err_code': 1,
                'err_msg': '文件路径不能为空',
                'payload': {}
            }), 400
        
        filepath = data['filepath']
        
        # 查询文档信息
        document_info = TgDocumentInfo.query.filter_by(filepath=filepath).first()
        
        if not document_info:
            return jsonify({
                'err_code': 1,
                'err_msg': '文档信息不存在',
                'payload': {}
            }), 404
        
        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': {
                'filepath': filepath,
                'chat_id': document_info.chat_id,
                'message_id': document_info.message_id,
                'filename_origin': document_info.filename_origin,
                'file_ext_name': document_info.file_ext_name,
                'mime_type': document_info.mime_type,
                'file_size': document_info.file_size,
                'file_hash': document_info.file_hash,
                'video_thumb_path': document_info.video_thumb_path
            }
        })
        
    except Exception as e:
        logger.error(f'根据路径获取文档信息失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


@api.route('/tg/user/stats/<user_id>', methods=['GET'])
def get_user_stats(user_id):
    """获取用户聊天统计数据"""
    try:
        # 验证参数
        if not user_id:
            return jsonify({
                'err_code': 1,
                'err_msg': '用户ID不能为空',
                'payload': {}
            }), 400
        
        # 查询用户的聊天统计数据
        stats_query = db.session.query(
            func.count(TgGroupChatHistory.id).label('total_messages'),
            func.min(TgGroupChatHistory.postal_time).label('first_message_time'),
            func.max(TgGroupChatHistory.postal_time).label('last_message_time')
        ).filter(TgGroupChatHistory.user_id == user_id)
        
        result = stats_query.first()
        
        # 查询用户所在的群组及其消息统计
        group_stats_query = db.session.query(
            TgGroupChatHistory.chat_id,
            func.count(TgGroupChatHistory.id).label('message_count'),
            func.max(TgGroupChatHistory.postal_time).label('last_active_time')
        ).filter(TgGroupChatHistory.user_id == user_id)\
         .group_by(TgGroupChatHistory.chat_id)\
         .order_by(func.max(TgGroupChatHistory.postal_time).desc())
        
        group_results = group_stats_query.all()
        
        # 获取群组信息
        chat_ids = [gr.chat_id for gr in group_results]
        groups = {}
        if chat_ids:
            group_info = TgGroup.query.filter(
                TgGroup.chat_id.in_(chat_ids),
                TgGroup.status == TgGroup.StatusType.JOIN_SUCCESS
            ).all()
            groups = {g.chat_id: g for g in group_info}
        
        # 构建群组列表
        user_groups = []
        for gr in group_results:
            group = groups.get(gr.chat_id)
            if group:
                user_groups.append({
                    'chat_id': gr.chat_id,
                    'title': group.title,
                    'name': group.name,
                    'message_count': gr.message_count,
                    'last_active_time': gr.last_active_time.strftime('%Y-%m-%d %H:%M:%S') if gr.last_active_time else ''
                })
        
        # 格式化统计数据
        stats = {
            'total_messages': result.total_messages or 0,
            'first_message_time': result.first_message_time.strftime('%Y-%m-%d %H:%M:%S') if result.first_message_time else '',
            'last_message_time': result.last_message_time.strftime('%Y-%m-%d %H:%M:%S') if result.last_message_time else '',
            'groups': user_groups
        }
        
        return jsonify({
            'err_code': 0,
            'err_msg': '',
            'payload': stats
        })
        
    except Exception as e:
        logger.error(f'获取用户统计数据失败: {e}')
        return jsonify({
            'err_code': 1,
            'err_msg': str(e),
            'payload': {}
        }), 500


def get_enhanced_document_info(chat_id, message_id, fallback_path=None, fallback_ext=None):
    """
    获取增强的文档信息
    优先使用数据库中的详细信息，如果没有则使用备用信息
    """
    try:
        # 查询数据库中的详细文档信息
        document_info = TgDocumentInfo.query.filter_by(
            chat_id=str(chat_id),
            message_id=str(message_id)
        ).first()
        
        if document_info and document_info.filename_origin:
            # 使用数据库中的详细信息
            return {
                'path': document_info.filepath or fallback_path or '',
                'filename_origin': document_info.filename_origin,
                'ext': document_info.file_ext_name or fallback_ext or '',
                'mime_type': document_info.mime_type or '',
                'file_hash': document_info.file_hash or '',
                'file_size': document_info.file_size or 0,
                'video_thumb_path': document_info.video_thumb_path or '',
                'has_detailed_info': True
            }
        elif fallback_path:
            # 使用备用信息
            doc_ext = fallback_ext or ''
            if not doc_ext and fallback_path:
                path_parts = fallback_path.split('.')
                if len(path_parts) > 1:
                    doc_ext = path_parts[-1]
            
            mime_type = get_mime_type_from_path(fallback_path, doc_ext)
            
            return {
                'path': fallback_path,
                'filename_origin': '',
                'ext': doc_ext,
                'mime_type': mime_type,
                'file_hash': '',
                'file_size': 0,
                'video_thumb_path': '',
                'has_detailed_info': False
            }
        else:
            return None
            
    except Exception as e:
        logger.error(f'获取增强文档信息失败: {e}')
        return None
